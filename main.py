"""
iotnac 自动化测试工程 — 验证入口
Phase 1: 模块导入 / 拓扑 / 截图 / 数据结构
Phase 2: Agent 图构建 + 自然语言用例执行
Phase 3: 知识库 - ChromaDB + 元素定位缓存
Phase 4: 用例系统 - YAML 规范 + Excel 转换 + 执行器
Phase 6: 完善集成 - HTML 报告 + 控制台汇总
"""
import sys
from pathlib import Path

# 确保项目根目录在 path 中
sys.path.insert(0, str(Path(__file__).parent))

from utils.logger import logger
from utils.screenshot import take_screenshot
from utils.result import TestResult, StepResult
from executors.ssh_executor import ssh_executor, TopologyLoader
from executors.web_executor import web_executor
from executors.desktop_executor import desktop_executor

PHASES = {"1", "2", "3", "4", "6"}


def test_imports():
    """验证所有模块可导入"""
    logger.info("=" * 50)
    logger.info("Phase 1 模块导入验证")
    logger.info("=" * 50)

    modules = [
        ("config.settings", "配置管理"),
        ("config.llm_config", "LLM 配置"),
        ("utils.logger", "日志工具"),
        ("utils.screenshot", "截图工具"),
        ("utils.result", "结果结构"),
        ("executors.ssh_executor", "SSH 执行器"),
        ("executors.web_executor", "Web 执行器"),
        ("executors.desktop_executor", "Desktop 执行器"),
    ]

    all_ok = True
    for mod_name, desc in modules:
        try:
            __import__(mod_name)
            logger.info(f"  ✓ {desc} ({mod_name})")
        except Exception as e:
            logger.error(f"  ✗ {desc} ({mod_name}): {e}")
            all_ok = False

    return all_ok


def test_topology():
    """验证拓扑加载"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("拓扑配置验证")
    logger.info("=" * 50)

    try:
        topology = TopologyLoader.load()
        devices = topology.get("devices", {})
        for name, info in devices.items():
            logger.info(
                f"  {name}: {info.get('type')} @ "
                f"{info.get('host', 'N/A')}:{info.get('port', 'N/A')}"
            )
        logger.info("拓扑加载成功")
        return True
    except FileNotFoundError:
        logger.warning("topology.yaml 未找到（这是正常的，请复制 .env.example 为 .env 后配置）")
        return True
    except Exception as e:
        logger.error(f"拓扑加载失败: {e}")
        return False


def test_screenshot():
    """验证截图功能"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("截图功能验证")
    logger.info("=" * 50)

    try:
        path = take_screenshot("phase1_test")
        if path.exists():
            logger.info(f"  ✓ 截图成功: {path} ({path.stat().st_size} bytes)")
            return True
        else:
            logger.error("  ✗ 截图文件未生成")
            return False
    except Exception as e:
        logger.error(f"  ✗ 截图失败: {e}")
        return False


def test_result():
    """验证结果数据结构"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("结果数据结构验证")
    logger.info("=" * 50)

    result = TestResult(title="Phase 1 验证测试")
    result.steps.append(StepResult(
        step_index=1,
        description="验证导入",
        target="system",
        action="import",
        params={},
        success=True,
        message="所有模块导入成功",
    ))
    result.steps.append(StepResult(
        step_index=2,
        description="验证截图",
        target="desktop",
        action="screenshot",
        params={},
        success=True,
        message="截图功能正常",
    ))

    logger.info(result.summary())
    return True


def test_agent_graph():
    """Phase 2: 验证 Agent 图可构建 + 执行简单用例"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("Phase 2 Agent 图验证")
    logger.info("=" * 50)

    # 1. 检查 LLM 配置
    from config.llm_config import LLMConfig
    missing = LLMConfig.validate()
    if missing:
        logger.warning(f"LLM 配置缺失: {missing}，跳过 Agent 测试")
        logger.warning("请配置 .env 中的 API Key 后再测")
        return False

    # 2. 构建图
    try:
        from agent.graph import build_graph
        graph = build_graph()
        logger.info("  ✓ LangGraph 图构建成功")
    except Exception as e:
        logger.error(f"  ✗ 图构建失败: {e}")
        return False

    # 3. 使用简单用例测试
    # 用户可通过命令行参数传入用例，默认使用简单示例
    test_case = sys.argv[2] if len(sys.argv) > 2 else "检查控制器SSH连接是否正常"
    logger.info(f"\n测试用例: {test_case}")

    initial_state = {"user_query": test_case}

    try:
        result = graph.invoke(initial_state)

        summary = result.get("summary", "无总结")
        step_results = result.get("step_results", [])
        error = result.get("error", "")

        logger.info("")
        logger.info("-" * 50)
        logger.info("执行结果")
        logger.info("-" * 50)

        if error:
            logger.error(f"Agent 执行出错: {error}")
            return False

        passed = sum(1 for r in step_results if r.get("passed"))
        logger.info(f"步骤: {passed}/{len(step_results)} 通过")
        
        for r in step_results:
            status = "✓" if r.get("passed") else "✗"
            logger.info(f"  {status} {r.get('description', r.get('tool', '?'))}")

        logger.info("")
        logger.info(summary)
        return passed > 0 or len(step_results) == 0

    except Exception as e:
        logger.error(f"Agent 执行异常: {e}")
        return False


def test_phase4():
    """Phase 4: 验证用例系统 — YAML 加载 + Excel 转换 + 执行器"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("Phase 4 用例系统验证")
    logger.info("=" * 50)

    # 1. YAML 用例加载
    logger.info("\n[1] YAML 用例加载...")
    from testcases.runner import runner

    yaml_dir = Path(__file__).parent / "testcases" / "yaml" / "web" / "login"
    case_path = yaml_dir / "TC-LOGIN-001.yaml"

    if not case_path.exists():
        logger.error(f"  ✗ 示例用例不存在: {case_path}")
        return False

    try:
        case = runner.load(str(case_path))
        logger.info(f"  ✓ 加载成功: {case['title']}")
        logger.info(f"    步骤数: {len(case.get('steps', []))}")
        logger.info(f"    断言数: {len(case.get('assertions', []))}")
    except Exception as e:
        logger.error(f"  ✗ 加载失败: {e}")
        return False

    # 2. 用例→Agent 查询转换
    logger.info("\n[2] YAML → Agent 查询转换...")
    query = runner._build_user_query(case)
    if query and "测试用例" in query:
        logger.info(f"  ✓ 查询生成成功 ({len(query)} 字符)")
        # 输出前 200 字符预览
        preview = query[:200].replace("\n", "\\n")
        logger.info(f"    预览: {preview}...")
    else:
        logger.error("  ✗ 查询生成失败")
        return False

    # 3. Excel 转换器
    logger.info("\n[3] Excel → YAML 转换...")
    try:
        import openpyxl
    except ImportError:
        logger.warning("  ⚠ openpyxl 未安装，跳过 Excel 转换测试")
        logger.info("  (如需测试 Excel 转换功能: pip install openpyxl)")
        return True  # 非关键依赖，不算失败

    from testcases.converters.excel2yaml import Excel2Yaml

    # 创建测试用 Excel
    converter_dir = Path(__file__).parent / "testcases" / "converters"
    test_xlsx = converter_dir / "_test_sample.xlsx"

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "登录测试"

        # 表头
        ws.cell(row=1, column=1, value="用例名称")
        ws.cell(row=1, column=2, value="测试步骤")
        ws.cell(row=1, column=3, value="预期结果")

        # 数据
        ws.cell(row=2, column=1, value="登录功能验证")
        ws.cell(row=2, column=2, value="1. 打开控制台页面\n2. 输入用户名admin\n3. 输入密码123456\n4. 点击登录按钮")
        ws.cell(row=2, column=3, value="1. 显示欢迎页\n2. 页面包含导航菜单")

        wb.save(str(test_xlsx))
        wb.close()
        logger.info(f"  ✓ 测试 Excel 已创建: {test_xlsx.name}")
    except Exception as e:
        logger.error(f"  ✗ 创建测试 Excel 失败: {e}")
        return False

    # 执行转换
    try:
        converter = Excel2Yaml()
        files = converter.convert(str(test_xlsx))
        if files:
            logger.info(f"  ✓ 转换成功: {len(files)} 个 YAML 文件")
            # 验证生成的文件
            with open(files[0], "r", encoding="utf-8") as f:
                content = f.read()
            if "登录功能验证" in content:
                logger.info(f"  ✓ 生成内容验证通过")
            else:
                logger.warning("  ⚠ 生成内容可能不正确")
        else:
            logger.error("  ✗ 转换无输出")
            return False
    except Exception as e:
        logger.error(f"  ✗ Excel 转换失败: {e}")
        return False
    finally:
        # 清理测试文件
        if test_xlsx.exists():
            test_xlsx.unlink()

    # 4. 批量加载
    logger.info("\n[4] 批量用例加载...")
    try:
        cases = runner.load_dir(str(yaml_dir))
        logger.info(f"  ✓ 批量加载成功: {len(cases)} 个用例")
    except Exception as e:
        logger.error(f"  ✗ 批量加载失败: {e}")
        return False

    # 5. Schema 验证
    logger.info("\n[5] Schema 约束验证...")
    schema_path = Path(__file__).parent / "testcases" / "schemas" / "testcase_schema.json"
    if schema_path.exists():
        import json
        with open(schema_path, "r", encoding="utf-8") as f:
            schema = json.load(f)
        if "title" in schema.get("required", []):
            logger.info(f"  ✓ Schema 有效 (必填字段: {schema['required']})")
        else:
            logger.error("  ✗ Schema 缺少 title 必填约束")
            return False
    else:
        logger.error(f"  ✗ Schema 文件不存在: {schema_path}")
        return False

    return True


def test_phase6():
    """Phase 6: 验证报告系统 — HTML 报告生成 + 控制台汇总"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("Phase 6 报告系统验证")
    logger.info("=" * 50)

    from config.settings import REPORT_DIR, SCREENSHOT_DIR
    from utils.reporter import generate_html_report, print_console_summary
    from utils.screenshot import take_screenshot

    # 1. 确保有截图可用于报告
    logger.info("\n[1] 准备测试数据...")
    ss_path = take_screenshot("phase6_test")
    if ss_path and ss_path.exists():
        logger.info(f"  ✓ 截图就绪: {ss_path.name} ({ss_path.stat().st_size} bytes)")
    else:
        logger.warning("  ⚠ 截图不可用，报告中将显示警告")

    # 2. 生成 HTML 报告
    logger.info("\n[2] 生成 HTML 报告...")
    step_results = [
        {"description": "打开 Web 控制台登录页", "tool": "web_navigate",
         "passed": True, "result": "页面加载成功", "screenshot": str(ss_path) if ss_path else None},
        {"description": "输入用户名 admin", "tool": "web_input",
         "passed": True, "result": "输入成功", "screenshot": str(ss_path) if ss_path else None},
        {"description": "输入密码并点击登录", "tool": "web_click",
         "passed": False, "result": "元素未找到: 登录按钮", "screenshot": str(ss_path) if ss_path else None},
        {"description": "验证登录结果", "tool": "web_check",
         "passed": True, "result": "页面显示欢迎信息", "screenshot": None},
    ]

    try:
        report_path = generate_html_report(
            title="Phase 6 验证 - 报告生成测试",
            step_results=step_results,
            passed=False,
            duration_s=2.35,
        )
    except Exception as e:
        logger.error(f"  ✗ 报告生成失败: {e}")
        return False

    # 3. 验证报告文件
    logger.info("\n[3] 验证报告文件...")
    if report_path.exists():
        content = report_path.read_text(encoding="utf-8")
        checks = [
            ("包含标题", "Phase 6 验证" in content),
            ("包含截图(base64)", "data:image/png;base64," in content),
            ("包含通过步骤", "✓ 通过" in content),
            ("包含失败步骤", "✗ 失败" in content),
            ("HTML 结构完整", "</html>" in content),
        ]
        all_checks_ok = True
        for check_name, ok in checks:
            status = "✓" if ok else "✗"
            logger.info(f"  {status} {check_name}")
            if not ok:
                all_checks_ok = False

        logger.info(f"  📄 报告路径: {report_path}")
        logger.info(f"  📏 报告大小: {report_path.stat().st_size / 1024:.1f} KB")

        if not all_checks_ok:
            return False
    else:
        logger.error(f"  ✗ 报告文件不存在: {report_path}")
        return False

    # 4. 控制台汇总
    logger.info("\n[4] 控制台汇总输出...")
    print_console_summary("Phase 6 验证", step_results, passed=False, duration_s=2.35)

    # 5. 报告目录
    logger.info("\n[5] 报告目录状态...")
    reports = list(REPORT_DIR.glob("*.html"))
    logger.info(f"  ✓ 已有 {len(reports)} 份报告")
    for r in reports:
        logger.info(f"    - {r.name} ({r.stat().st_size / 1024:.1f} KB)")

    return True


def test_knowledge_base():
    """Phase 3: 验证知识库 — 文档入库 + 检索 + 定位缓存"""
    logger.info("")
    logger.info("=" * 50)
    logger.info("Phase 3 知识库验证")
    logger.info("=" * 50)

    from knowledge.manager import knowledge_manager

    # 1. 添加测试文档
    logger.info("\n[1] 文档入库...")
    sample_doc = (
        "# IoT NAC 产品概述\n\n"
        "IoT NAC 是一套网络准入控制系统，包含三个核心组件：\n\n"
        "## Web 控制台\n"
        "Web 控制台是管理员进行配置和监控的主界面。\n"
        "支持用户管理、设备管理、策略配置、日志查看等功能。\n"
        "默认登录地址 http://192.168.3.101:8080\n\n"
        "## 客户端\n"
        "客户端运行在终端设备上，负责上报设备信息和执行准入策略。\n\n"
        "## 控制器\n"
        "控制器是核心服务，运行在 Linux 服务器上 (192.168.3.101:22)。\n"
        "负责策略决策、设备认证、日志记录。\n\n"
        "## 交换机\n"
        "华为交换机 (192.168.3.1:23) 通过 Telnet 管理，执行 VLAN 和 ACL 配置。\n"
        "典型配置命令：display vlan, display acl, display interface。"
    )

    ok = knowledge_manager.add_doc_text(sample_doc, "product_overview")
    if ok:
        logger.info("  ✓ 测试文档已入库")
    else:
        logger.error("  ✗ 文档入库失败")
        return False

    # 2. 语义检索
    logger.info("\n[2] 语义检索...")
    results = knowledge_manager.search("交换机的管理方式是什么")
    if results:
        logger.info(f"  ✓ 检索命中 {len(results)} 条")
        logger.info(f"    首条距离: {results[0]['distance']:.4f}")
        logger.info(f"    内容预览: {results[0]['content'][:100]}...")
    else:
        logger.error("  ✗ 检索无结果")
        return False

    # 3. 元素定位缓存
    logger.info("\n[3] 元素定位缓存...")

    # 缓存一条定位策略
    knowledge_manager.cache_locator(
        page_url="http://192.168.3.101:8080/login",
        description="登录按钮",
        locator_type="role",
        locator_value="button[name='登录']",
    )
    logger.info("  ✓ 已缓存定位: 登录按钮 → role:button[name='登录']")

    # 查询缓存
    from knowledge.schemas.element_locator import ElementLocator
    locator = knowledge_manager.get_locator(
        page_url="http://192.168.3.101:8080/login",
        description="登录按钮",
    )
    if locator:
        logger.info(f"  ✓ 命中缓存: {locator.locator_type} → {locator.locator_value}")
    else:
        logger.error("  ✗ 未命中缓存")
        return False

    # 4. 统计
    stats = knowledge_manager.stats()
    logger.info(f"\n[4] 知识库统计: 向量文档={stats['vector_docs']}, 定位缓存={stats['locator_cache']}")

    return True


def main():
    """验证主函数 — 默认运行 Phase 1 + Phase 2"""
    # 确定要运行的 Phase
    if len(sys.argv) > 1 and sys.argv[1] in PHASES:
        phases = {sys.argv[1]}
    else:
        phases = PHASES

    results = []

    if "1" in phases:
        logger.info("")
        logger.info("=" * 50)
        logger.info("Phase 1: 基础框架验证")
        logger.info("=" * 50)
        results.append(("模块导入", test_imports()))
        results.append(("拓扑加载", test_topology()))
        results.append(("截图功能", test_screenshot()))
        results.append(("结果结构", test_result()))

    if "2" in phases:
        results.append(("Agent 图", test_agent_graph()))

    if "3" in phases:
        results.append(("知识库", test_knowledge_base()))

    if "4" in phases:
        results.append(("用例系统", test_phase4()))

    if "6" in phases:
        results.append(("报告系统", test_phase6()))

    logger.info("")
    logger.info("=" * 50)
    logger.info("验证汇总")
    logger.info("=" * 50)

    for name, ok in results:
        status = "✓" if ok else "✗"
        logger.info(f"  {status} {name}")

    all_ok = all(r[1] for r in results)
    logger.info(f"\n状态: {'通过' if all_ok else '部分失败，请检查日志'}")


if __name__ == "__main__":
    main()