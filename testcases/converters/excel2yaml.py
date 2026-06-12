"""Excel 测试用例 → YAML 格式转换器

V1 简化版：将 Excel 中的用例数据转换为 YAML 用例文件。
支持从测试平台导出的 Excel 格式。

Excel 列映射规则：
  - 用例名称 / 标题 → title
  - 测试步骤 / 操作步骤 → steps（按行拆分）
  - 预期结果 → assertions（按行拆分）

未识别格式的步骤标记为 # TODO: 需人工确认
"""
import re
import sys
from pathlib import Path
from typing import Optional

import yaml

# 确保项目根目录在 path 中
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from utils.logger import logger
from knowledge.manager import knowledge_manager


class Excel2Yaml:
    """Excel → YAML 转换器"""

    # 列名关键词映射（支持中英文模糊匹配）
    COLUMN_MAP = {
        "title": ["用例名", "标题", "用例名称", "title", "用例标题", "测试用例"],
        "description": ["描述", "说明", "description", "用例描述", "前置条件"],
        "steps": ["步骤", "测试步骤", "操作步骤", "steps", "测试内容"],
        "assertions": ["预期", "预期结果", "断言", "assertions", "期望结果", "验证点"],
    }

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Args:
            output_dir: YAML 输出目录，默认 testcases/yaml/converted/
        """
        if output_dir is None:
            output_dir = Path(__file__).parent.parent / "yaml" / "converted"
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ── 公开接口 ─────────────────────────────────

    def convert(self, excel_path: str) -> list[Path]:
        """转换整个 Excel 文件，返回生成的 YAML 文件路径列表"""
        try:
            import openpyxl
        except ImportError:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            return []

        path = Path(excel_path)
        if not path.exists():
            logger.error(f"Excel 文件不存在: {excel_path}")
            return []

        wb = openpyxl.load_workbook(path, data_only=True)
        output_files = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"处理工作表: {sheet_name} ({ws.max_row} 行)")

            column_indices = self._detect_columns(ws)
            if not column_indices.get("title"):
                logger.warning(f"  未识别到 标题列，跳过")
                continue

            # 按行解析用例（每行一个用例）
            for row in range(2, ws.max_row + 1):
                case = self._parse_row(ws, row, column_indices)
                if case and case.get("title"):
                    yaml_path = self._save_case(case)
                    output_files.append(yaml_path)
                    logger.info(f"  ✓ 已生成: {yaml_path.name}")

        wb.close()
        logger.info(f"转换完成: {len(output_files)} 个用例 → {self.output_dir}")
        return output_files

    # ── 内部方法 ─────────────────────────────────

    def _detect_columns(self, ws) -> dict[str, int]:
        """自动识别第一行列名，返回 {field: column_index}"""
        indices = {}
        # 先收集所有表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=col).value or "").strip().lower()
            if val:
                headers[col] = val

        # 匹配已知列
        for field, keywords in self.COLUMN_MAP.items():
            for col, val in headers.items():
                if any(kw in val for kw in keywords):
                    indices[field] = col
                    break

        logger.info(f"  识别列: {indices}")
        return indices

    def _parse_row(self, ws, row: int, col_idx: dict) -> Optional[dict]:
        """解析单行用例"""
        title = self._cell(ws, row, col_idx.get("title"))
        if not title:
            return None

        case = {
            "title": title,
            "steps": self._parse_steps(self._cell(ws, row, col_idx.get("steps"))),
            "assertions": self._parse_assertions(self._cell(ws, row, col_idx.get("assertions"))),
        }

        desc = self._cell(ws, row, col_idx.get("description"))
        if desc:
            case["description"] = desc

        return case

    def _parse_steps(self, raw: str) -> list[dict]:
        """将步骤文本拆分为结构化步骤列表"""
        if not raw:
            return []

        steps = []
        # 按数字序号或换行拆分步骤
        raw = raw.strip()
        # 尝试按 "1." "2." 等数字序号拆分
        numbered = re.split(r"\n(?=\d+[\.\、\)）])", raw)
        if len(numbered) > 1:
            parts = numbered
        else:
            # 按换行拆分
            parts = [s.strip() for s in raw.split("\n") if s.strip()]

        for part in parts:
            part = re.sub(r"^\d+[\.\、\)\uff09]\s*", "", part).strip()
            if not part:
                continue
        
            step = self._classify_step(part)
            # 用知识库补全步骤信息
            step = self._enrich_step_with_knowledge(step)
            steps.append(step)
        
        return steps

    def _classify_step(self, text: str) -> dict:
        """根据文本内容推断 target 和 action"""
        text_lower = text.lower()

        step = {"description": text}

        # 判断目标
        if any(kw in text_lower for kw in ["ssh", "控制器", "交换机", "switch", "controller"]):
            step["target"] = "ssh"
        elif any(kw in text_lower for kw in ["客户端", "桌面", "desktop", "窗口"]):
            step["target"] = "desktop"
        else:
            step["target"] = "web"

        # 判断操作
        if any(kw in text_lower for kw in ["打开", "访问", "导航", "navigate", "进入"]):
            step["action"] = "navigate"
        elif any(kw in text_lower for kw in ["输入", "填写", "键入", "input", "enter"]):
            step["action"] = "input"
        elif any(kw in text_lower for kw in ["点击", "click", "选择", "select"]):
            step["action"] = "click"
        elif any(kw in text_lower for kw in ["执行", "运行", "execute", "命令"]):
            step["action"] = "execute"
        elif any(kw in text_lower for kw in ["截图", "screenshot"]):
            step["action"] = "screenshot"
        elif any(kw in text_lower for kw in ["等待", "wait"]):
            step["action"] = "wait"
        elif any(kw in text_lower for kw in ["检查", "验证", "确认", "check", "assert"]):
            step["action"] = "check"
        else:
            step["action"] = "click"

        step["params"] = {}
        step["_auto_classified"] = True  # 标记为自动识别

        return step

    def _enrich_step_with_knowledge(self, step: dict) -> dict:
        """用知识库补全步骤信息（如元素描述、表单字段等）"""
        desc = step.get("description", "")
        if not desc:
            return step

        try:
            results = knowledge_manager.search(desc, top_k=2)
            if not results:
                return step

            # 只取距离较近的结果（相似度高的）
            best = results[0]
            if best.get("distance", 1.0) > 0.3:
                return step

            content = best.get("content", "")
            source = best.get("metadata", {}).get("source", "")

            # 根据操作类型补全参数
            action = step.get("action", "")
            params = step.get("params", {})

            if action == "input" and "element" not in params:
                # 从知识库提取输入框描述
                element_desc = self._extract_element_from_knowledge(desc, content)
                if element_desc:
                    params["element"] = element_desc
                    step["_knowledge_source"] = source

            elif action == "click" and "element" not in params:
                # 从知识库提取按钮/链接描述
                element_desc = self._extract_element_from_knowledge(desc, content)
                if element_desc:
                    params["element"] = element_desc
                    step["_knowledge_source"] = source

            elif action == "navigate" and "url" not in params:
                # 从知识库提取 URL
                url = self._extract_url_from_knowledge(content)
                if url:
                    params["url"] = url
                    step["_knowledge_source"] = source

            # 补充知识库中发现的额外上下文
            context = self._extract_context_hints(content)
            if context:
                step["_knowledge_hints"] = context

            step["params"] = params

        except Exception as e:
            logger.debug(f"知识库补全异常: {e}")

        return step

    @staticmethod
    def _extract_element_from_knowledge(step_desc: str, content: str) -> str:
        """从知识库内容中提取与步骤描述匹配的元素名称"""
        import re as _re
        # 查找知识库中包含步骤关键词的行
        keywords = _re.findall(r'[\u4e00-\u9fff]+', step_desc)
        if not keywords:
            return ""

        # 在知识库内容中查找包含这些关键词的元素描述
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
            # 匹配包含输入框/按钮/菜单等元素的描述
            if any(kw in line for kw in keywords):
                if any(hint in line for hint in ['输入框', '按钮', '菜单', '链接', '复选框', '下拉框', '表单']):
                    # 提取元素描述（截取合适的长度）
                    element = line.strip('*').strip('`').strip()
                    if len(element) > 50:
                        element = element[:50]
                    return element

        return ""

    @staticmethod
    def _extract_url_from_knowledge(content: str) -> str:
        """从知识库内容中提取 URL"""
        import re as _re
        urls = _re.findall(r'https?://[\w\.:\-/]+', content)
        return urls[0] if urls else ""

    @staticmethod
    def _extract_context_hints(content: str) -> list[str]:
        """从知识库中提取有用的上下文提示（如前置条件、注意事项）"""
        hints = []
        for line in content.split('\n'):
            line = line.strip()
            if any(hint in line for hint in ['前置条件', '注意', '必须', '需要先', '确保']):
                if len(line) < 200:
                    hints.append(line)
        return hints[:3]  # 最多3条提示

    def _parse_assertions(self, raw: str) -> list[str]:
        """将预期结果文本拆分为断言列表"""
        if not raw:
            return []

        raw = raw.strip()
        # 按序号或换行拆分
        numbered = re.split(r"\n(?=\d+[\.\、\)）])", raw)
        if len(numbered) > 1:
            return [re.sub(r"^\d+[\.\、\)）]\s*", "", a).strip() for a in numbered if a.strip()]

        return [s.strip() for s in raw.split("\n") if s.strip()]

    def _save_case(self, case: dict) -> Path:
        """保存用例为 YAML 文件"""
        # 生成文件名（标题转英文标识）
        filename = self._title_to_filename(case["title"])
        yaml_path = self.output_dir / f"{filename}.yaml"

        # 避免重名
        counter = 1
        while yaml_path.exists():
            yaml_path = self.output_dir / f"{filename}_{counter}.yaml"
            counter += 1

        content = yaml.dump(case, allow_unicode=True, default_flow_style=False, sort_keys=False)
        yaml_path.write_text(content, encoding="utf-8")
        return yaml_path

    @staticmethod
    def _cell(ws, row: int, col) -> str:
        """安全读取单元格值"""
        if col is None:
            return ""
        val = ws.cell(row=row, column=col).value
        return str(val).strip() if val else ""

    @staticmethod
    def _title_to_filename(title: str) -> str:
        """标题 → 文件名（中文保留拼音首字母策略，V1 使用清理后的标题）"""
        # 移除特殊字符，保留中英文和数字
        cleaned = re.sub(r"[^\w\u4e00-\u9fff\-]", "_", title)
        # 去除连续下划线
        cleaned = re.sub(r"_+", "_", cleaned).strip("_")
        return cleaned or "test_case"


def main():
    """命令行入口"""
    if len(sys.argv) < 2:
        print("用法: python excel2yaml.py <excel文件路径> [输出目录]")
        sys.exit(1)

    excel_path = sys.argv[1]
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    converter = Excel2Yaml(output_dir)
    files = converter.convert(excel_path)

    if files:
        print(f"\n成功生成 {len(files)} 个 YAML 文件:")
        for f in files:
            print(f"  {f}")
    else:
        print("\n未生成任何 YAML 文件，请检查 Excel 格式")


if __name__ == "__main__":
    main()
